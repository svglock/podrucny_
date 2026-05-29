import json
import tempfile
import xml.etree.ElementTree as ET

import openpyxl
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.http import FileResponse
from django.http import HttpResponse
from django.shortcuts import render, redirect
from documents.models import Invoice, CommercialOffer, ShippingDocument
from documents.pdf import generate_invoice_pdf
from inventory.models import StockMovement
from logistics.models import Delivery
from orders.models import Order, OrderItem
from warehouse.models import Product

from .forms import (
    CreateDeliveryForm,
    LoginForm,
    CreateOrderForm,
    ExcelImportForm,
    XMLImportForm
)


@login_required
def dashboard_view(request):
    total_revenue = Order.objects.aggregate(
        total=Sum('total_price')
    )['total'] or 0

    low_stock = Product.objects.filter(
        quantity__lt=10
    ).count()
    new_orders = Order.objects.filter(status='new').count()
    confirmed_orders = Order.objects.filter(status='confirmed').count()
    delivered_orders = Order.objects.filter(status='delivered').count()
    low_stock_products = Product.objects.filter(quantity__lt=10)[:5]
    recent_orders = Order.objects.order_by(
        '-created_at'
    )[:5]
    latest_movements = (
        StockMovement.objects
        .select_related('product')
        .order_by('-created_at')[:5]
    )
    # Данные для графика выручки по месяцам
    revenue_by_month = (
        Order.objects
        .filter(status='delivered')
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(total=Sum('total_price'))
        .order_by('-month')[:6]
    )

    # Формируем списки в хронологическом порядке
    labels_list = [d['month'].strftime('%b %Y') for d in reversed(list(revenue_by_month))]
    values_list = [float(d['total'] or 0) for d in reversed(list(revenue_by_month))]

    # Превращаем в JSON-строки
    chart_labels_json = json.dumps(labels_list)
    chart_values_json = json.dumps(values_list)
    context = {
        'products_count': Product.objects.count(),
        'orders_count': Order.objects.count(),
        'deliveries_count': Delivery.objects.count(),
        'invoices_count': Invoice.objects.count(),
        'total_revenue': total_revenue,
        'low_stock': low_stock,
        'new_orders': new_orders,
        'confirmed_orders': confirmed_orders,
        'delivered_orders': delivered_orders,
        'low_stock_products': low_stock_products,
        'recent_orders': recent_orders,
        'latest_movements': latest_movements,
        'chart_labels_json': chart_labels_json,
        'chart_values_json': chart_values_json,
    }
    return render(request, 'dashboard/dashboard.html', context)


@login_required
def products_view(request):
    products = Product.objects.all()

    search = request.GET.get(
        'search'
    )

    if search:
        products = products.filter(
            title__icontains=search
        )

    return render(
        request,
        'dashboard/products.html',
        {
            'products': products
        }
    )


@login_required
def orders_view(request):
    orders = Order.objects.all()

    if request.user.role == 'customer':
        orders = orders.filter(
            customer=request.user
        )

    search = request.GET.get('search')

    if search:
        orders = orders.filter(
            id=search
        )

    status = request.GET.get('status')

    if status:
        orders = orders.filter(
            status=status
        )

    return render(
        request,
        'dashboard/orders.html',
        {
            'orders': orders
        }
    )


@login_required
def deliveries_view(request):
    if request.user.role == 'courier':
        deliveries = Delivery.objects.filter(courier=request.user)
    elif request.user.role in ['manager', 'admin']:
        deliveries = Delivery.objects.all()
    else:
        deliveries = Delivery.objects.none()
    return render(request, 'dashboard/deliveries.html', {
        'deliveries': deliveries
    })


@login_required
def documents_view(request):
    return render(request, 'dashboard/documents.html', {
        'invoices': Invoice.objects.all(),
        'offers': CommercialOffer.objects.all(),
        'shipping': ShippingDocument.objects.all(),
    })


def login_view(request):
    form = LoginForm(request.POST or None)
    if request.method == 'POST':
        user = authenticate(
            username=form.data.get('username'),
            password=form.data.get('password')
        )
        if user:
            login(request, user)
            return redirect('/')
    return render(request, 'dashboard/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('/login/')


@login_required
def change_order_status(request, order_id, status):
    order = Order.objects.get(id=order_id)
    order.status = status
    order.save()
    return redirect('/orders/')


@login_required
def change_delivery_status(
        request,
        delivery_id,
        status
):
    delivery = Delivery.objects.get(
        id=delivery_id
    )

    delivery.delivery_status = status
    delivery.save()

    return redirect('/deliveries/')


@login_required
def invoice_pdf(request, invoice_id):
    invoice = Invoice.objects.get(id=invoice_id)
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    generate_invoice_pdf(invoice, temp_file.name)
    return FileResponse(
        open(temp_file.name, 'rb'),
        as_attachment=True,
        filename=f'invoice_{invoice.id}.pdf'
    )


@login_required
def create_order_view(request):
    form = CreateOrderForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        product = form.cleaned_data['product']
        quantity = form.cleaned_data['quantity']
        order = Order.objects.create(
            customer_company=form.cleaned_data[
                'customer_company'
            ],
            customer=request.user,
            status='new',
            total_price=0
        )
        item_price = product.final_price * quantity
        OrderItem.objects.create(
            order=order,
            product=product,
            quantity=quantity,
            price=item_price
        )
        order.total_price = item_price
        order.save()
        return redirect('/orders/')
    return render(request, 'dashboard/create_order.html', {'form': form})


@login_required
def import_excel_view(request):
    form = ExcelImportForm(
        request.POST or None,
        request.FILES or None
    )

    if request.method == 'POST' and form.is_valid():

        workbook = openpyxl.load_workbook(
            request.FILES['file']
        )

        sheet = workbook.active

        for row in sheet.iter_rows(
                min_row=2,
                values_only=True
        ):
            Product.objects.create(
                supplier=request.user,
                title=row[0],
                description=row[1],
                price=row[2],
                quantity=row[3],
                category='other'
            )

        return redirect('/products/')

    return render(
        request,
        'dashboard/import_excel.html',
        {
            'form': form
        }
    )


@login_required
def import_xml_view(request):
    form = XMLImportForm(
        request.POST or None,
        request.FILES or None
    )

    if request.method == 'POST' and form.is_valid():

        tree = ET.parse(
            request.FILES['file']
        )

        root = tree.getroot()

        for product in root.findall('product'):
            Product.objects.create(
                supplier=request.user,
                title=product.find(
                    'title'
                ).text,

                description=product.find(
                    'description'
                ).text,

                price=product.find(
                    'price'
                ).text,

                quantity=product.find(
                    'quantity'
                ).text,

                category='other'
            )

        return redirect('/products/')

    return render(
        request,
        'dashboard/import_xml.html',
        {
            'form': form
        }
    )


@login_required
def create_delivery_view(request):
    form = CreateDeliveryForm(
        request.POST or None
    )

    if request.method == 'POST' and form.is_valid():
        Delivery.objects.create(
            order=form.cleaned_data['order'],
            courier=form.cleaned_data['courier'],
            start_address=form.cleaned_data['start_address'],
            end_address=form.cleaned_data['end_address'],
            distance_km=form.cleaned_data['distance_km'],
            route_duration_minutes=form.cleaned_data['route_duration_minutes']
        )

        return redirect('/deliveries/')

    return render(
        request,
        'dashboard/create_delivery.html',
        {
            'form': form
        }
    )


@login_required
def shipping_pdf(
        request,
        shipping_id
):
    document = ShippingDocument.objects.get(
        id=shipping_id
    )

    temp_file = tempfile.NamedTemporaryFile(
        delete=False,
        suffix='.pdf'
    )

    from documents.pdf import (
        generate_shipping_pdf
    )

    generate_shipping_pdf(
        document,
        temp_file.name
    )

    return FileResponse(
        open(temp_file.name, 'rb'),
        as_attachment=True,
        filename=f'shipping_{shipping_id}.pdf'
    )


@login_required
def export_orders_excel(request):
    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.append([
        'ID',
        'Customer',
        'Status',
        'Total'
    ])

    for order in Order.objects.all():
        sheet.append([
            order.id,
            str(order.customer),
            order.status,
            float(order.total_price)
        ])

    response = HttpResponse(
        content_type=
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response[
        'Content-Disposition'
    ] = (
        'attachment; '
        'filename=orders.xlsx'
    )

    workbook.save(response)

    return response


@login_required
def stock_movements_view(request):
    movements = StockMovement.objects.select_related(
        'product',
        'created_by'
    ).order_by('-created_at')

    movement_type = request.GET.get(
        'type'
    )

    if movement_type:
        movements = movements.filter(
            movement_type=movement_type
        )

    return render(
        request,
        'dashboard/stock_movements.html',
        {
            'movements': movements
        }
    )


@login_required
def profile_view(request):
    return render(
        request,
        'dashboard/profile.html'
    )


from documents.models import ShippingDocument


@login_required
def shipping_document_view(request, order_id):
    # Пытаемся получить существующий документ, если его нет – создаём с минимальными данными
    document, created = ShippingDocument.objects.get_or_create(
        order_id=order_id,
        defaults={
            'route': 'Не назначен',
            'cargo_description': 'Электронное оборудование',
            'cargo_weight': 12.5,
            'places_count': 1
        }
    )
    return render(request, 'dashboard/shipping_document.html', {'document': document})
